import matplotlib.pyplot as plt
import pdfkit
import pathlib
from utils import Utils
from jinja2 import Environment, FileSystemLoader
from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Alignment, Side
from openpyxl.utils import get_column_letter


class Report:
    def __init__(self):
        self.workbook = Workbook()

    def generate_excel(
            self,
            all_salary_level: Dict[str, int],
            all_vacancies_count: Dict[str, int],
            salary_level: Dict[str, int],
            vacancies_count: Dict[str, int],
            by_city_level: Dict[str, float],
            vacancies_part: Dict[str, float]
    ) -> None:
        self.workbook.active.title = "Статистика по годам"
        self.workbook.create_sheet("Статистика по городам", 1)
        self.__make_year(
            self.workbook["Статистика по годам"],
            all_salary_level,
            all_vacancies_count,
            salary_level,
            vacancies_count
        )
        self.__make_city(
            self.workbook["Статистика по городам"],
            Utils.slice_dict(by_city_level, 10),
            Utils.slice_dict(vacancies_part, 10)
        )
        self.__stylizing_xlsx()

        self.workbook.save("report.xlsx")

    def generate_image(
            self,
            vacancy_name: str,
            all_salary_level: Dict[str, int],
            all_vacancies_count: Dict[str, int],
            salary_level: Dict[str, int],
            vacancies_count: Dict[str, int],
            by_city_level: Dict[str, float],
            vacancies_part: Dict[str, float]
    ) -> None:
        fig = plt.figure()
        salary_year = fig.add_subplot(2, 2, 1)
        vacancies_year = fig.add_subplot(2, 2, 2)
        salary_city = fig.add_subplot(2, 2, 3)
        part_city = fig.add_subplot(2, 2, 4)

        salary_year.bar(
            [int(e) - 0.2 for e in all_salary_level.keys()],
            list(all_salary_level.values()),
            width=0.4,
            label="средняя з/п"
        )
        salary_year.bar(
            [int(e) + 0.2 for e in salary_level.keys()],
            list(salary_level.values()),
            width=0.4,
            label=f"з/п {vacancy_name}"
        )

        vacancies_year.bar(
            [int(e) - 0.2 for e in all_vacancies_count.keys()],
            list(all_vacancies_count.values()),
            width=0.4,
            label="Количество вакансий"
        )
        vacancies_year.bar(
            [int(e) + 0.2 for e in vacancies_count.keys()],
            list(vacancies_count.values()),
            width=0.4,
            label=f"Количество вакансий {vacancy_name}"
        )

        by_city_level_sliced = dict(reversed(Utils.slice_dict(by_city_level, 10).items()))
        salary_city.barh(
            list(by_city_level_sliced.keys()),
            list(by_city_level_sliced.values())
        )

        vacancies_part_sliced = Utils.slice_dict(vacancies_part, 10)
        other_part = sum(Utils.dict_difference(vacancies_part_sliced, vacancies_part).values())
        part_city.pie(
            [e * 100 for e in vacancies_part_sliced.values()] + [other_part * 100],
            labels=list(vacancies_part_sliced.keys()) + ["Другие"],
            textprops={'fontsize': 6}
        )

        self.__format_histograms(
            salary_year,
            vacancies_year,
            salary_city,
            part_city
        )
        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(
            self,
            vacancy_name
    ) -> None:
        env = Environment(loader=FileSystemLoader("."))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({
            "vacancy_name": vacancy_name,
            "img": pathlib.Path("graph.png").resolve().__str__(),
            "workbook": self.workbook
        })

        config = pdfkit.configuration(wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, "report.pdf", configuration=config, options={"enable-local-file-access": True})

    @staticmethod
    def __format_histograms(
            salary_year,
            vacancies_year,
            salary_city,
            part_city
    ) -> None:
        salary_year.set_title("Уровень зарплат по годам")
        salary_year.grid(axis='y')
        salary_year.tick_params(labelsize=8)
        salary_year.set_xticks(salary_year.get_xticks(), salary_year.get_xticklabels(), rotation=90)
        salary_year.legend(fontsize=8)

        vacancies_year.set_title("Количество вакансий по годам")
        vacancies_year.grid(axis='y')
        vacancies_year.tick_params(labelsize=8)
        vacancies_year.set_xticks(vacancies_year.get_xticks(), vacancies_year.get_xticklabels(), rotation=90)
        vacancies_year.legend(fontsize=8)

        salary_city.grid(axis='x')
        salary_city.set_yticklabels(
            labels=[Utils.wrap_text(e.get_text(), [" ", "-"]) for e in salary_city.get_yticklabels()],
            fontsize=6
        )
        salary_city.tick_params(axis="x", labelsize=8)
        salary_city.set_title("Уровень зарплат по городам")

        part_city.set_title("Доля вакансий по городам")

    def __stylizing_xlsx(self) -> None:
        bd = Side(style='thin', color="000000")
        for sheet in self.workbook:
            column_widths = []
            for row_count, row in enumerate(sheet.iter_rows()):
                for i, cell in enumerate(row):
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    if row_count == 0:
                        cell.font = Font(bold=True)
                    else:
                        cell.alignment = Alignment(horizontal="right")
                    if len(column_widths) > i:
                        if len(str(cell.value)) > column_widths[i]:
                            column_widths[i] = len(str(cell.value))
                    else:
                        column_widths += [len(str(cell.value))]

            for i, column_width in enumerate(column_widths, 1):
                sheet.column_dimensions[get_column_letter(i)].width = column_width + 4

        for percent_cell in self.workbook["Статистика по городам"]["E"]:
            percent_cell.number_format = "0.00%"

    @staticmethod
    def __make_year(
            worksheet,
            all_salary_level: Dict[str, int],
            all_vacancies_count: Dict[str, int],
            salary_level: Dict[str, int],
            vacancies_count: Dict[str, int]
    ) -> None:
        worksheet.append([
            "Год",
            "Динамика уровня зарплат по годам",
            "Динамика количества вакансий по годам",
            "Динамика уровня зарплат по годам для выбранной профессии",
            "Динамика количества вакансий по годам для выбранной профессии"
        ])
        for key in all_salary_level.keys():
            worksheet.append([
                key,
                all_salary_level[key],
                all_vacancies_count[key],
                salary_level[key],
                vacancies_count[key]
            ])

    @staticmethod
    def __make_city(
            worksheet,
            by_city_level: Dict[str, float],
            vacancies_part: Dict[str, float]
    ) -> None:
        worksheet.append([
            "Город",
            "Уровень зарплат по городам (в порядке убывания)",
            "",
            "Город",
            "Доля вакансий по городам (в порядке убывания)"
        ])
        level_keys = list(by_city_level.keys())
        part_keys = list(vacancies_part.keys())
        for i in range(len(level_keys)):
            level_key = level_keys[i]
            part_key = part_keys[i]
            worksheet.append([
                level_key,
                by_city_level[level_key],
                "",
                part_key,
                vacancies_part[part_key]
            ])