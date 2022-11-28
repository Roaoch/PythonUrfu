import csv
import codecs
import re
import matplotlib.pyplot as plt
from datetime import datetime
from typing import List, Generator, Dict
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Alignment, Side
from openpyxl.utils import get_column_letter


class OutOfDataError(BaseException):
    pass


class Utils:
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    @staticmethod
    def add_to_or_update(dictionary: Dict[any, any], key: any, value: any) -> None:
        if dictionary.__contains__(key):
            dictionary[key] += value
        else:
            dictionary.update({key: value})

    @staticmethod
    def slice_dict(dictionary: dict, end: int):
        result = {}
        i = 0
        for key, value in dictionary.items():
            i += 1
            if i <= end:
                result.update({key: value})
            else:
                break
        return result

    @staticmethod
    def dict_difference(sliced_dict: Dict[str, float], full_dict: Dict[str, float]) -> Dict[str, float]:
        return {key: full_dict[key] for key in full_dict.keys() if key not in sliced_dict.keys()}

    @staticmethod
    def wrap_text(text: str, separators: List[str]):
        result = text
        for sep in separators:
            if text.__contains__(sep):
                result = "\n".join(text.split(sep, 1))
                break
        return result


class Salary:
    def __init__(self, salary_property: List[str]):
        self.salary_from = float(salary_property[0])
        self.salary_to = float(salary_property[1])
        if len(salary_property) == 4:
            self.salary_gross = salary_property[2]
            self.salary_currency = salary_property[3]
        else:
            self.salary_currency = salary_property[2]

    def get_salary(self) -> float:
        return (self.salary_from + self.salary_to) / 2 * Utils.currency_to_rub[self.salary_currency]


class Vacancy:
    def __init__(self, property_list: List[any], headline: List[str]):
        if len(headline) == 12:
            self.name = property_list[0]
            self.description = property_list[1]
            self.key_skills = property_list[2] if type(property_list[2]) == list else [property_list[2]]
            self.experience_id = property_list[3]
            self.premium = property_list[4]
            self.employer_name = property_list[5]
            self.salary = Salary(property_list[6:10])
            self.area_name = property_list[10]
            self.published_at = datetime.strptime(property_list[11], "%Y-%m-%dT%H:%M:%S%z")
        elif len(headline) == 6:
            self.name = property_list[0]
            self.salary = Salary(property_list[1:4])
            self.area_name = property_list[4]
            self.published_at = datetime.strptime(property_list[5], "%Y-%m-%dT%H:%M:%S%z")


class DataSet:
    __CLEANER = re.compile('<.*?>')
    __SPACE_CLEANER = re.compile('(\s\s+)|(\xa0)')

    def __init__(self, file_name: str):
        self.file_name = file_name
        self.headline = []
        self.vacancies_reader = self.__clean_properties(
            vacancies=self.__reader(file_name=file_name)
        )

    def __reader(self, file_name: str) -> csv.reader:
        reader = csv.reader(codecs.open(file_name, "r", "utf_8_sig"), delimiter=',')
        self.headline = reader.__next__()
        return reader

    def __validate(self, element: List[str]) -> bool:
        if len(element) != len(self.headline):
            return False
        for i in range(len(element)):
            if element[i] == '':
                return False
        return True

    def __clean_properties(self, vacancies: csv.reader) -> Generator[Vacancy, None, None]:
        i = 0
        key_skills_index = self.headline.index("key_skills") if "key_skills_index" in self.headline else -1
        for vacancy in vacancies:
            i += 1
            clean_vacancy = []
            if self.__validate(element=vacancy):
                for merit_index in range(len(vacancy)):
                    if merit_index == key_skills_index:
                        temp_property = vacancy[merit_index].split('\n')
                        for i in range(len(temp_property)):
                            temp_property[i] = re.sub(self.__CLEANER, '', temp_property[i])
                            temp_property[i] = re.sub(self.__SPACE_CLEANER, ' ', temp_property[i]).strip()
                    else:
                        temp_property = re.sub(self.__CLEANER, '', vacancy[merit_index])
                        temp_property = re.sub(self.__SPACE_CLEANER, ' ', temp_property).strip()
                    clean_vacancy.append(temp_property)
                yield Vacancy(clean_vacancy, self.headline)
        if i == 0:
            raise OutOfDataError


class InputConnect:
    def __init__(
            self,
            filename,
            filter_parameter
    ):
        self.__vacancies = DataSet(file_name=filename.strip())
        self.filter_parameter = filter_parameter.strip()
        self.all_salary_level = {}
        self.all_vacancies_count = {}
        self.salary_level = {}
        self.vacancies_count = {}
        self.by_city_level = {}
        self.vacancies_part = {}
        self.__get_statistics()

    def __get_statistics(self) -> None:
        temp_by_city_count = {}
        temp_by_city_level = {}
        temp_all_by_city_count = {}
        for vacancy in self.__vacancies.vacancies_reader:
            vacancy_year = vacancy.published_at.year
            vacancy_salary = vacancy.salary.get_salary()
            Utils.add_to_or_update(
                self.all_salary_level,
                vacancy_year,
                vacancy_salary
            )
            Utils.add_to_or_update(
                self.all_vacancies_count,
                vacancy_year,
                1
            )
            Utils.add_to_or_update(
                temp_by_city_level,
                vacancy.area_name,
                vacancy_salary
            )
            Utils.add_to_or_update(
                temp_all_by_city_count,
                vacancy.area_name,
                1
            )
            if self.filter_parameter in vacancy.name:
                Utils.add_to_or_update(
                    self.salary_level,
                    vacancy_year,
                    vacancy_salary
                )
                Utils.add_to_or_update(
                    self.vacancies_count,
                    vacancy_year,
                    1
                )
                Utils.add_to_or_update(
                    temp_by_city_count,
                    vacancy.area_name,
                    1
                )
        self.all_salary_level = {
            key: int(value / self.all_vacancies_count[key]) for key, value in self.all_salary_level.items()
        }
        self.salary_level = {
            key: int(value / self.vacancies_count[key]) for key, value in self.salary_level.items()
        }

        if len(self.salary_level) == 0:
            self.salary_level = {key: 0 for key in self.all_vacancies_count.keys()}
            self.vacancies_count = {key: 0 for key in self.all_vacancies_count.keys()}

        self.vacancies_part = self.__get_vacancy_part(temp_all_by_city_count)
        self.by_city_level = dict(sorted(
            [(key, int(temp_by_city_level[key] / temp_all_by_city_count[key])) for key in self.vacancies_part.keys()],
            key=lambda e: (e[1], -len(e[0])),
            reverse=True
        ))

    def print_self(self) -> None:
        print(f"Динамика уровня зарплат по годам: {self.all_salary_level}")
        print(f"Динамика количества вакансий по годам: {self.all_vacancies_count}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {self.salary_level}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {self.vacancies_count}")
        print(f"Уровень зарплат по городам (в порядке убывания): {Utils.slice_dict(self.by_city_level, 10)}")
        print(f"Доля вакансий по городам (в порядке убывания): {Utils.slice_dict(self.vacancies_part, 10)}")

    @staticmethod
    def __get_vacancy_part(by_city_count: Dict[str, int]):
        result = []
        all_vacancy_count = sum(by_city_count.values())
        for key, value in by_city_count.items():
            part = value / all_vacancy_count
            if part >= 0.01:
                result.append((key, float("{:.4f}".format(part))))
        return dict(sorted(
            result,
            key=lambda e: e[1],
            reverse=True
        ))


class Report:
    def __init__(self):
        self.workbook = Workbook()

    def generate_excel(
            self,
            all_salary_level: Dict[str, int],
            all_vacancies_count: Dict[str, int],
            salary_level: Dict[str, int],
            vacancies_count: Dict[str, int],
            by_city_level: Dict[str, float],
            vacancies_part: Dict[str, float]
    ) -> None:
        self.workbook.active.title = "Статистика по годам"
        self.workbook.create_sheet("Статистика по городам", 1)
        self.__make_year(
            self.workbook["Статистика по годам"],
            all_salary_level,
            all_vacancies_count,
            salary_level,
            vacancies_count
        )
        self.__make_city(
            self.workbook["Статистика по городам"],
            Utils.slice_dict(by_city_level, 10),
            Utils.slice_dict(vacancies_part, 10)
        )
        self.__stylizing_xlsx()

        self.workbook.save("report.xlsx")

    def generate_image(
            self,
            vacancy_name: str,
            all_salary_level: Dict[str, int],
            all_vacancies_count: Dict[str, int],
            salary_level: Dict[str, int],
            vacancies_count: Dict[str, int],
            by_city_level: Dict[str, float],
            vacancies_part: Dict[str, float]
    ) -> None:
        fig = plt.figure()
        salary_year = fig.add_subplot(2, 2, 1)
        vacancies_year = fig.add_subplot(2, 2, 2)
        salary_city = fig.add_subplot(2, 2, 3)
        part_city = fig.add_subplot(2, 2, 4)

        salary_year.bar(
            [int(e) - 0.2 for e in all_salary_level.keys()],
            list(all_salary_level.values()),
            width=0.4,
            label="средняя з/п"
        )
        salary_year.bar(
            [int(e) + 0.2 for e in salary_level.keys()],
            list(salary_level.values()),
            width=0.4,
            label=f"з/п {vacancy_name}"
        )

        vacancies_year.bar(
            [int(e) - 0.2 for e in all_vacancies_count.keys()],
            list(all_vacancies_count.values()),
            width=0.4,
            label="Количество вакансий"
        )
        vacancies_year.bar(
            [int(e) + 0.2 for e in vacancies_count.keys()],
            list(vacancies_count.values()),
            width=0.4,
            label=f"Количество вакансий {vacancy_name}"
        )

        by_city_level_sliced = dict(reversed(Utils.slice_dict(by_city_level, 10).items()))
        salary_city.barh(
            list(by_city_level_sliced.keys()),
            list(by_city_level_sliced.values())
        )

        vacancies_part_sliced = Utils.slice_dict(vacancies_part, 10)
        other_part = sum(Utils.dict_difference(vacancies_part_sliced, vacancies_part).values())
        part_city.pie(
            [e * 100 for e in vacancies_part_sliced.values()] + [other_part * 100],
            labels=list(vacancies_part_sliced.keys()) + ["Другие"],
            textprops={'fontsize': 6}
        )

        self.__format_histograms(
            salary_year,
            vacancies_year,
            salary_city,
            part_city
        )
        plt.show()

    @staticmethod
    def __format_histograms(
            salary_year,
            vacancies_year,
            salary_city,
            part_city
    ) -> None:
        salary_year.set_title("Уровень зарплат по годам")
        salary_year.grid(axis='y')
        salary_year.tick_params(labelsize=8)
        salary_year.set_xticks(salary_year.get_xticks(), salary_year.get_xticklabels(), rotation=90)
        salary_year.legend(fontsize=8)

        vacancies_year.set_title("Количество вакансий по годам")
        vacancies_year.grid(axis='y')
        vacancies_year.tick_params(labelsize=8)
        vacancies_year.set_xticks(vacancies_year.get_xticks(), vacancies_year.get_xticklabels(), rotation=90)
        vacancies_year.legend(fontsize=8)

        salary_city.grid(axis='x')
        salary_city.set_yticklabels(
            labels=[Utils.wrap_text(e.get_text(), [" ", "-"]) for e in salary_city.get_yticklabels()],
            fontsize=6
        )
        salary_city.tick_params(axis="x", labelsize=8)
        salary_city.set_title("Уровень зарплат по городам")

        part_city.set_title("Доля вакансий по городам")

    def __stylizing_xlsx(self) -> None:
        bd = Side(style='thin', color="000000")
        for sheet in self.workbook:
            column_widths = []
            for row_count, row in enumerate(sheet.iter_rows()):
                for i, cell in enumerate(row):
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    if row_count == 0:
                        cell.font = Font(bold=True)
                    else:
                        cell.alignment = Alignment(horizontal="right")
                    if len(column_widths) > i:
                        if len(str(cell.value)) > column_widths[i]:
                            column_widths[i] = len(str(cell.value))
                    else:
                        column_widths += [len(str(cell.value))]

            for i, column_width in enumerate(column_widths, 1):
                sheet.column_dimensions[get_column_letter(i)].width = column_width + 4

        for percent_cell in self.workbook["Статистика по городам"]["E"]:
            percent_cell.number_format = "0.00%"

    @staticmethod
    def __make_year(
            worksheet,
            all_salary_level: Dict[str, int],
            all_vacancies_count: Dict[str, int],
            salary_level: Dict[str, int],
            vacancies_count: Dict[str, int]
    ) -> None:
        worksheet.append([
            "Год",
            "Динамика уровня зарплат по годам",
            "Динамика количества вакансий по годам",
            "Динамика уровня зарплат по годам для выбранной профессии",
            "Динамика количества вакансий по годам для выбранной профессии"
        ])
        for key in all_salary_level.keys():
            worksheet.append([
                key,
                all_salary_level[key],
                all_vacancies_count[key],
                salary_level[key],
                vacancies_count[key]
            ])

    @staticmethod
    def __make_city(
            worksheet,
            by_city_level: Dict[str, float],
            vacancies_part: Dict[str, float]
    ) -> None:
        worksheet.append([
            "Город",
            "Уровень зарплат по городам (в порядке убывания)",
            "",
            "Город",
            "Доля вакансий по городам (в порядке убывания)"
        ])
        level_keys = list(by_city_level.keys())
        part_keys = list(vacancies_part.keys())
        for i in range(len(level_keys)):
            level_key = level_keys[i]
            part_key = part_keys[i]
            worksheet.append([
                level_key,
                by_city_level[level_key],
                "",
                part_key,
                vacancies_part[part_key]
            ])


try:
    input_connect = InputConnect(
        input("Введите название файла: "),
        input("Введите название профессии: ")
    )
    report = Report()

    report.generate_excel(
        input_connect.all_salary_level,
        input_connect.all_vacancies_count,
        input_connect.salary_level,
        input_connect.vacancies_count,
        input_connect.by_city_level,
        input_connect.vacancies_part
    )
    report.generate_image(
        input_connect.filter_parameter,
        input_connect.all_salary_level,
        input_connect.all_vacancies_count,
        input_connect.salary_level,
        input_connect.vacancies_count,
        input_connect.by_city_level,
        input_connect.vacancies_part
    )
    input_connect.print_self()
except StopIteration:
    print("Пустой файл")
except IOError:
    print("Формат ввода некорректен")
except KeyError:
    print("Параметр поиска некорректен")
except AssertionError:
    print("Ничего не найдено")
except OutOfDataError:
    print("Нет данных")
